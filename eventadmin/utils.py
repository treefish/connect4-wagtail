"""
Author..: David Apimerika
Created.: 24 May 2023
Updated.: 21 July 2023
          06 December 2023 (copied from connect4-django)

Purpose.:
"""

#import datetime
from datetime import date, time, datetime
import pytz
from dateutil.parser import parse
import itertools
from openpyxl import load_workbook, utils
from openpyxl.styles import Font, Fill, PatternFill
from copy import copy

# from django.db.models import Count
from django.contrib.auth import get_user_model

from events.models import ProjectPage, EventPage, EventType
from bookings.models import Booking, Attendance

from registration.models import FamilyMember
# from registration.stats import FamilyMemberStats
# from events.models import Project, EventType, Event, Booking, Attendance

User = get_user_model()

########################################################################################################################
#                                                                                                                      #
########################################################################################################################
def create_attendance_register_daily(event, filename):
    workbook = load_workbook(filename="data/templates/attendance_register_daily.xlsx")
    workbook.iso_dates = True
    event_ws = workbook["Day Register"]
    ws_date_format = "%d-%b"
    tz = pytz.timezone('Pacific/Auckland')  # 'Europe/London'
    start_date = event.start_date
    ws_name = start_date.astimezone(tz).strftime(ws_date_format)
    run_dt = datetime.now().astimezone(tz).replace(tzinfo=None)
    event_date = start_date.astimezone(tz).replace(tzinfo=None)
    print(f"  - Event: {event}\t{ws_name}")
    event_ws.title = ws_name
    event_ws["B1"] = event.title
    event_ws["B2"] = event_date
    event_ws["B3"] = event.week
    event_ws["S1"] = event.id
#    event_ws["S2"] = event.wp_event_id
    event_ws["S4"] = run_dt

    # Loop through booked attendees and copy their info into a row for each, starting at Row 6.
    row = 7
    current_family_name = ""
    bookings = Booking.objects.filter(event=event)
    attendance_list = (
        Attendance.objects
#        .select_related("family_member__family")
        .select_related("booking__family")
        .filter(booking__in=bookings)
        .order_by(
            "booking__family__last_name",
            "booking__family__first_name",
        )
    )
    for attendee in attendance_list:
        row += 1
        family_member = attendee.family_member
        registrant = f"{attendee.booking.family.first_name} {attendee.booking.family.last_name}"
        print(f"    - Attendee: {row - 5}\t{family_member}")
        event_ws[f"A{row}"] = family_member.first_name
        event_ws[f"B{row}"] = family_member.last_name
        event_ws[f"E{row}"] = family_member.type.capitalize()
        event_ws[f"H{row}"] = 1 if attendee.attended else ""
        event_ws[f"R{row}"] = family_member.id
        if family_member.type == "CHILD":
            event_ws[f"C{row}"] = 1 if family_member.childmore.fsm else ""
            event_ws[f"D{row}"] = family_member.childmore.sen_detail
            event_ws[f"F{row}"] = family_member.childmore.dob

        family_name = family_member.family.family_name
        if not current_family_name == family_name:
            event_ws[f"S{row}"] = registrant
            event_ws[f"S{row}"].fill = copy(event_ws["S1"].fill)
            event_ws[f"T{row}"] = family_name
            event_ws[f"T{row}"].fill = copy(event_ws["S1"].fill)
            current_family_name = family_name

    # Save spreadsheet
    workbook.save(filename=f"media/admin/{filename}")
    workbook.close()


########################################################################################################################
#                                                                                                                      #
########################################################################################################################
def import_attendance_register_daily(filename, event_id):
    """
    This function loads an Attendance Register Daily spreadsheet that has been downloaded from this website.
    It creates/updates bookings (Booking) and Attendance entries for Family Members/Attendees (Attendance).
    Check this spreadsheet has the right event_id.

    It is called from the Event Detail page.
    """
    # Get the Event
    try:
        event = EventPage.objects.get(id=event_id)
        print(f" - Found event: {event} ({event.event_type})")
    except EventPage.DoesNotExist:
        print(f" - No such event! Cannot proceed.")
        return

    print(f"+++++++++++ Sheet {filename} +++++++++++")
    ws_date_format = "%d-%b"
    wb = load_workbook(filename=filename, data_only=True)
    event_ws = wb.active
    print(f"Processing: {event.id}\t{event.title}")
    unfound_members_list = []
    bad_sen_detail = ['none', 'n/a', 'na', 'no', '-', 'fsm']   # Use to clean up SEN detail
    row = 8
    first_name = event_ws[f"A{row}"].value
    while isinstance(first_name, str):
        last_name = event_ws[f"B{row}"].value
        family_member_id = event_ws[f"R{row}"].value
        attended_value = event_ws[f"H{row}"].value
        attended = (
            True if (attended_value == 1) or (attended_value == "Present") else False
        )
        print(f"    - {row}\t{first_name} {last_name}\t {family_member_id}\t{attended}")
        try:
            # Find the Family member first, keyed to 'id'.
            family_member = FamilyMember.objects.get(id=family_member_id)
            print(
                f"      - Found family_member: {family_member.id}\t{family_member}"
            )
            # Check if names are the same. Could be a legit name change, especially last name, but both names different would suggest
            # the wp_member_id is not unique.
            if (family_member.first_name != first_name) or (
                family_member.last_name != last_name
            ):
                unfound_members_list.append(f"{family_member_id} - {first_name} {last_name}")
                print(
                    f"      - ! {family_member.first_name} {family_member.last_name} is NOT the same as {first_name} {last_name}"
                )
                row += 1
                first_name = event_ws[f"A{row}"].value
                continue

            # Find or Create the Booking for the Registrant/Organiser
            user = User.objects.get(family_member=family_member)
            print(f"      - Found family: {user.family_name}")
            booking, created = Booking.objects.get_or_create(family=user, event=event)
            if created:
                print(f"      - Creating new booking.")
            else:
                print(f"      - Booking already created.")

            # Find or Create the Attendance entry for the Person
            attendance, created = Attendance.objects.get_or_create(
                booking=booking,
                family_member=family_member,
            )
            attendance.attended = attended
            attendance.save()

            # Update other fields as required.
            # SEN - Col D - text of SEN details.
            # Could try to weed out incorrect ones - a list of No, None, n/a, etc
            #print(f"      - Updating SEN info.")
            if family_member.type == "CHILD":
                sen_detail = str(event_ws[f"D{row}"].value).strip() if event_ws[f"D{row}"].value else ""
                if len(sen_detail) == 0:
                    family_member.childmore.sen_detail = ""
                    family_member.childmore.sen_req = False
                elif sen_detail.lower() in bad_sen_detail:
                    print(f"      - Updating SEN info - removing {sen_detail}.")
                    family_member.childmore.sen_detail = ""
                    family_member.childmore.sen_req = False
                else:
                    family_member.childmore.sen_detail = sen_detail
                    family_member.childmore.sen_req = True
                family_member.childmore.save()

        except FamilyMember.DoesNotExist:
            unfound_members_list.append(f"{wp_member_id} - {first_name} {last_name}")
            print(f"  *** No such Family Member: {wp_member_id} - {first_name} {last_name}")

        row += 1
        first_name = event_ws[f"A{row}"].value

    wb.close()
    print(f"Unfound Family members: {len(unfound_members_list)}")
    for unfound in unfound_members_list:
        print(f"- {unfound}")


########################################################################################################################
#  Projects                                                                                                            #
########################################################################################################################
def create_attendance_register(project_id, filename):
    workbook = load_workbook(filename="data/templates/attendance_register.xlsx")
    workbook.iso_dates = True
    template_sheet = workbook["Day Register"]
    tz = pytz.timezone('Pacific/Auckland')
    ws_date_format = "%d-%b"

    # Get the Project
    try:
        project = ProjectPage.objects.get(id=project_id)
        print(f" - Found Project: {project}")
    except ProjectPage.DoesNotExist:
        print(f" - No such Project! Cannot proceed.")
        return

    # Just Family Events for now. May need to create a different function for Youth Events.
    event_type = EventType.objects.get(name="Family Fun Days")
    event_list = EventPage.objects.live().descendant_of(project).filter(event_type=event_type).order_by("-start_date")
    #events_list = EventPage.objects.filter(projectpage=project, event_type=event_type).order_by(
    #     "start_date"
    # )
    # For each Event in the Project
    summary_seq = 4  # Row in Summary worksheet to write summary info.
    for event in event_list:
        summary_seq += 1
        event_date = event.start_date.astimezone(tz).replace(tzinfo=None)
        # Copy the template worksheet to a new one for the event, naming it with the date of the event, e.g. 03-Apr
        ws_name = event_date.astimezone(tz).strftime(ws_date_format)
        print(f"  - Event: {event}\t{ws_name}")
        event_ws = workbook.copy_worksheet(template_sheet)
        event_ws.title = ws_name
        event_ws["B1"] = event.title
        event_ws["B2"] = event_date
        event_ws["B3"] = event.week

        # Loop through booked attendees and copy their info into a row for each, starting at Row 6.
        row = 7
        current_family_name = ""

# The block here is from daily. Compare!
        # bookings = Booking.objects.filter(event=event)
        # attendance_list = (
        #     Attendance.objects.select_related("family_member__family")
        #     .select_related("booking__family__registrant")
        #     .filter(booking__in=bookings)
        #     .order_by(
        #         "booking__family__registrant__last_name",
        #         "booking__family__registrant__first_name",
        #     )
        # )
        #

        for attendee in (
            Attendance.objects.filter(booking__event=event).select_related("family_member__family")
            .select_related("booking__family")
            .order_by("booking__family__last_name")
            .order_by("booking__family__first_name")
        ):
            row += 1
            family_member = attendee.family_member
            registrant = f"{attendee.booking.family.first_name} {attendee.booking.family.last_name}"
            print(f"    - Attendee: {row - 7}\t{family_member}")
            event_ws[f"A{row}"] = family_member.first_name
            event_ws[f"B{row}"] = family_member.last_name
            event_ws[f"E{row}"] = family_member.type.capitalize()
            event_ws[f"H{row}"] = 1 if attendee.attended else ""
            event_ws[f"Q{row}"] = family_member.id
            if family_member.type == "CHILD":
                event_ws[f"C{row}"] = 1 if family_member.childmore.fsm else ""
                event_ws[f"D{row}"] = (
                    family_member.childmore.sen_detail
                    if family_member.childmore.sen_req
                    else ""
                )
                event_ws[f"F{row}"] = family_member.childmore.dob

            family_name = family_member.family.family_name

            if not current_family_name == family_name:
                event_ws[f"R{row}"] = registrant
                event_ws[f"R{row}"].fill = copy(event_ws["R1"].fill)
                event_ws[f"S{row}"] = family_name
                event_ws[f"S{row}"].fill = copy(event_ws["R1"].fill)
                current_family_name = family_name

        # Add Summary info
        summary_ws = workbook["Summary"]
        # TODO: summary_ws[f"B1"] = event.parent.title
        summary_ws[f"A{summary_seq}"] = event.id
        summary_ws[f"B{summary_seq}"] = event.title
        summary_ws[f"C{summary_seq}"] = event_ws["B3"].value
        summary_ws[f"D{summary_seq}"] = event_ws["B2"].value
        summary_ws[f"E{summary_seq}"] = event_ws[
            "B2"
        ].value
        summary_ws[f"F{summary_seq}"] = f"={utils.quote_sheetname(event_ws.title)}!B5"
        summary_ws[f"G{summary_seq}"] = f"={utils.quote_sheetname(event_ws.title)}!H5"
        summary_ws[f"H{summary_seq}"] = f"={utils.quote_sheetname(event_ws.title)}!C5"
        summary_ws[f"I{summary_seq}"] = f"={utils.quote_sheetname(event_ws.title)}!D5"
        summary_ws[f"J{summary_seq}"] = f"={utils.quote_sheetname(event_ws.title)}!I5"
        summary_ws[f"K{summary_seq}"] = f"={utils.quote_sheetname(event_ws.title)}!J5"
        summary_ws[f"L{summary_seq}"] = f"={utils.quote_sheetname(event_ws.title)}!K5"
        summary_ws[f"M{summary_seq}"] = f"={utils.quote_sheetname(event_ws.title)}!L5"
        summary_ws[f"N{summary_seq}"] = f"={utils.quote_sheetname(event_ws.title)}!M5"
        summary_ws[f"O{summary_seq}"] = f"={utils.quote_sheetname(event_ws.title)}!N5"
        summary_ws[f"P{summary_seq}"] = f"={utils.quote_sheetname(event_ws.title)}!O5"
        summary_ws[f"Q{summary_seq}"] = f"={utils.quote_sheetname(event_ws.title)}!P5"
        summary_ws[f"R{summary_seq}"] = f"={utils.quote_sheetname(event_ws.title)}!U5"

    # Save spreadsheet
    workbook.remove(workbook["Day Register"])
    workbook.save(filename=f"media/admin/{filename}")

